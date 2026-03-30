"""
Master Data Template Generator
Generates a single Excel workbook with all 8 data sheets + README,
complete with column definitions, sample data, formatting, and validation dropdowns.
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── Style constants ──────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
DEF_FONT = Font(name="Calibri", italic=True, color="555555", size=10)
DEF_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
NEW_COL_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
NEW_COL_FILL = PatternFill(start_color="00897b", end_color="00897b", fill_type="solid")
NEW_DEF_FILL = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
SAMPLE_FONT = Font(name="Calibri", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

TAB_RED = "FF1744"
TAB_ORANGE = "FF9100"
TAB_GREEN = "00C853"
TAB_BLUE = "2979FF"
TAB_PURPLE = "7B1FA2"


# ── Sheet definitions ────────────────────────────────────────────────────────

SHEETS = [
    {
        "name": "stores",
        "priority": "REQUIRED",
        "tab_color": TAB_RED,
        "frequency": "One-time setup (update when stores open/close)",
        "columns": [
            {"col": "store_id",      "type": "Text",    "definition": "Unique store identifier. Use same ID in ALL other files. Format: XX-NNN (e.g. RH-001)", "example": "RH-001", "new": False},
            {"col": "name",          "type": "Text",    "definition": "Store display name for dashboards and reports", "example": "Hypermarket Hlaing", "new": False},
            {"col": "sector",        "type": "Text",    "definition": "Business sector: Retail / F&B / Distribution / Property", "example": "Retail", "new": False, "dropdown": "Retail,F&B,Distribution,Property"},
            {"col": "channel",       "type": "Text",    "definition": "Channel within sector (e.g. Hypermarket, Supermarket, Convenience, Bakery, Restaurant, Beverage, QSR, Warehouse, Office, Mall, Industrial Park)", "example": "Hypermarket", "new": False, "dropdown": "Hypermarket,Supermarket,Convenience,Bakery,Restaurant,Beverage,QSR,Warehouse,Office,Mall,Industrial Park"},
            {"col": "township",      "type": "Text",    "definition": "Physical location/township. Used for blackout zone clustering and heatmap", "example": "Hlaing", "new": False},
            {"col": "has_solar",     "type": "Boolean", "definition": "Does this site have solar panels installed? TRUE or FALSE", "example": "TRUE", "new": False, "dropdown": "TRUE,FALSE"},
            {"col": "generator_kw",  "type": "Number",  "definition": "Backup generator capacity in kilowatts. Used to calculate expected diesel consumption", "example": "250", "new": False},
            {"col": "cold_chain",    "type": "Boolean", "definition": "Does this site have cold storage (dairy, frozen, fresh)? Enables spoilage prediction", "example": "TRUE", "new": False, "dropdown": "TRUE,FALSE"},
        ],
        "sample_rows": [
            ["RH-001", "Hypermarket Hlaing", "Retail", "Hypermarket", "Hlaing", "TRUE", 250, "TRUE"],
            ["FB-001", "Bakery Junction Square", "F&B", "Bakery", "Kamayut", "FALSE", 80, "TRUE"],
            ["DW-001", "Central Warehouse", "Distribution", "Warehouse", "Mingalardon", "TRUE", 400, "TRUE"],
        ],
    },
    {
        "name": "daily_energy",
        "priority": "REQUIRED",
        "tab_color": TAB_RED,
        "frequency": "Daily — submit by 8:00 PM each day",
        "columns": [
            {"col": "date",                   "type": "Date",    "definition": "Record date in YYYY-MM-DD format", "example": "2026-03-15", "new": False},
            {"col": "store_id",               "type": "Text",    "definition": "Must match a store_id in the stores sheet", "example": "RH-001", "new": False},
            {"col": "blackout_hours",         "type": "Number",  "definition": "Total hours without grid power today (0-24). Measure from grid outage start to restore.", "example": "5.5", "new": False},
            {"col": "generator_hours",        "type": "Number",  "definition": "Total hours generator ran today. RULE: generator_hours + grid_hours must be <= 24", "example": "4.0", "new": False},
            {"col": "grid_hours",             "type": "Number",  "definition": "Hours on grid power = 24 - blackout_hours (auto-calculate or enter manually)", "example": "18.5", "new": False},
            {"col": "diesel_consumed_liters", "type": "Number",  "definition": "Diesel burned by generator today in liters. Read from meter or estimate from hours x rate.", "example": "32.0", "new": False},
            {"col": "diesel_cost_mmk",        "type": "Number",  "definition": "Cost of diesel consumed today = liters x price per liter (MMK)", "example": "96000", "new": False},
            {"col": "grid_cost_mmk",          "type": "Number",  "definition": "Grid electricity cost for the day in MMK", "example": "46250", "new": False},
            {"col": "solar_kwh",              "type": "Number",  "definition": "Solar energy generated today in kWh. Enter 0 if no solar. Get from inverter portal.", "example": "85.3", "new": False},
            {"col": "total_energy_cost_mmk",  "type": "Number",  "definition": "Total = diesel_cost_mmk + grid_cost_mmk (MMK)", "example": "142250", "new": False},
            {"col": "operating_mode_actual",  "type": "Text",    "definition": "NEW — What mode did the store ACTUALLY operate in today? Used to track AI recommendation adoption.", "example": "FULL", "new": True, "dropdown": "FULL,SELECTIVE,CRITICAL,CLOSED"},
            {"col": "operating_mode_planned", "type": "Text",    "definition": "NEW — What did the AI COMMANDER recommend? Auto-filled by system if available.", "example": "FULL", "new": True, "dropdown": "FULL,SELECTIVE,CRITICAL,CLOSED"},
            {"col": "submitted_by",           "type": "Text",    "definition": "NEW — Name of the data champion who submitted this record", "example": "U Aung Ko", "new": True},
            {"col": "submitted_at",           "type": "DateTime","definition": "NEW — Timestamp when data was submitted. Used for submission compliance tracking.", "example": "2026-03-15 19:30:00", "new": True},
        ],
        "sample_rows": [
            ["2026-03-15", "RH-001", 5.5, 4.0, 18.5, 32.0, 96000, 46250, 85.3, 142250, "FULL", "FULL", "U Aung Ko", "2026-03-15 19:30:00"],
            ["2026-03-15", "FB-001", 8.0, 7.5, 16.0, 18.0, 54000, 20000, 0, 74000, "SELECTIVE", "SELECTIVE", "Ma Thida", "2026-03-15 18:45:00"],
            ["2026-03-15", "DW-001", 3.0, 3.0, 21.0, 48.0, 144000, 52500, 120.5, 196500, "FULL", "FULL", "Ko Zaw", "2026-03-15 20:10:00"],
        ],
    },
    {
        "name": "diesel_prices",
        "priority": "REQUIRED",
        "tab_color": TAB_RED,
        "frequency": "Daily — one row per day (market level, not per store)",
        "columns": [
            {"col": "date",                   "type": "Date",    "definition": "Price date in YYYY-MM-DD format", "example": "2026-03-15", "new": False},
            {"col": "diesel_price_mmk",       "type": "Number",  "definition": "Diesel price per liter in Myanmar Kyat from your primary supplier", "example": "4850", "new": False},
            {"col": "fx_usd_mmk",             "type": "Number",  "definition": "USD to MMK exchange rate. Correlated with diesel price.", "example": "3520", "new": False},
            {"col": "brent_oil_usd",          "type": "Number",  "definition": "Brent crude oil price in USD per barrel. External signal for forecast.", "example": "88.50", "new": False},
            {"col": "price_change_pct",       "type": "Number",  "definition": "% change from yesterday's price. Positive = price went up.", "example": "2.3", "new": False},
            {"col": "regional_shortage_flag",  "type": "Text",    "definition": "NEW — Is there a reported fuel shortage in the region? Y = shortage reported, N = normal supply", "example": "N", "new": True, "dropdown": "Y,N"},
            {"col": "supplier_source",         "type": "Text",    "definition": "NEW — Which supplier provided this price quote. Tracks supplier reliability.", "example": "Myanmar Petroleum", "new": True},
        ],
        "sample_rows": [
            ["2026-03-15", 4850, 3520, 88.50, 2.3, "N", "Myanmar Petroleum"],
            ["2026-03-16", 4920, 3535, 89.10, 1.4, "N", "Myanmar Petroleum"],
            ["2026-03-17", 5100, 3580, 91.20, 3.7, "Y", "Shwe Taung Fuel"],
        ],
    },
    {
        "name": "diesel_inventory",
        "priority": "RECOMMENDED",
        "tab_color": TAB_ORANGE,
        "frequency": "Daily — end-of-day tank reading per store",
        "columns": [
            {"col": "date",                   "type": "Date",    "definition": "Record date in YYYY-MM-DD format", "example": "2026-03-15", "new": False},
            {"col": "store_id",               "type": "Text",    "definition": "Must match a store_id in the stores sheet", "example": "RH-001", "new": False},
            {"col": "diesel_stock_liters",    "type": "Number",  "definition": "Diesel remaining in tank at end of day (liters). Dipstick or gauge reading.", "example": "320.5", "new": False},
            {"col": "diesel_purchased_liters","type": "Number",  "definition": "Diesel delivered/purchased today (liters). Enter 0 if no delivery.", "example": "200", "new": False},
            {"col": "tank_capacity_liters",   "type": "Number",  "definition": "Maximum fuel tank capacity (liters). Usually stays same.", "example": "500", "new": False},
            {"col": "supplier_lead_time_days","type": "Number",  "definition": "Current estimated delivery lead time in days from your supplier", "example": "1.5", "new": False},
            {"col": "days_of_coverage",       "type": "Number",  "definition": "How many days current stock will last = stock / avg daily consumption", "example": "4.2", "new": False},
        ],
        "sample_rows": [
            ["2026-03-15", "RH-001", 320.5, 200, 500, 1.5, 4.2],
            ["2026-03-15", "FB-001", 85.0, 0, 150, 2.0, 2.1],
            ["2026-03-15", "DW-001", 650.0, 400, 1000, 1.0, 5.8],
        ],
    },
    {
        "name": "store_sales",
        "priority": "RECOMMENDED",
        "tab_color": TAB_ORANGE,
        "frequency": "Daily — hourly breakdown per store (export from POS system)",
        "columns": [
            {"col": "date",             "type": "Date",    "definition": "Sales date in YYYY-MM-DD format", "example": "2026-03-15", "new": False},
            {"col": "hour",             "type": "Number",  "definition": "Hour of day (6-21 for operating hours 6am to 10pm)", "example": "10", "new": False},
            {"col": "store_id",         "type": "Text",    "definition": "Must match a store_id in the stores sheet", "example": "RH-001", "new": False},
            {"col": "sales_mmk",        "type": "Number",  "definition": "Total revenue for this hour in MMK. Export from POS.", "example": "1250000", "new": False},
            {"col": "gross_margin_mmk", "type": "Number",  "definition": "Gross margin = sales - cost of goods sold (MMK)", "example": "275000", "new": False},
            {"col": "transactions",     "type": "Number",  "definition": "Number of customer transactions this hour", "example": "85", "new": False},
            {"col": "labour_cost_mmk",  "type": "Number",  "definition": "NEW — Total staff labour cost allocated to this hour (MMK). Used for EBITDA per hour calculation. If unknown, leave blank — system uses channel average.", "example": "45000", "new": True},
        ],
        "sample_rows": [
            ["2026-03-15", 10, "RH-001", 1250000, 275000, 85, 45000],
            ["2026-03-15", 11, "RH-001", 1480000, 325600, 102, 45000],
            ["2026-03-15", 12, "RH-001", 1820000, 400400, 135, 52000],
        ],
    },
    {
        "name": "fx_rates",
        "priority": "RECOMMENDED",
        "tab_color": TAB_ORANGE,
        "frequency": "Daily — one row per day (from bank or central bank)",
        "columns": [
            {"col": "date",              "type": "Date",    "definition": "Rate date in YYYY-MM-DD format", "example": "2026-03-15", "new": False},
            {"col": "usd_mmk",           "type": "Number",  "definition": "US Dollar to Myanmar Kyat rate", "example": "3520", "new": False},
            {"col": "eur_mmk",           "type": "Number",  "definition": "Euro to Myanmar Kyat rate", "example": "3802", "new": False},
            {"col": "sgd_mmk",           "type": "Number",  "definition": "Singapore Dollar to Myanmar Kyat rate", "example": "2605", "new": False},
            {"col": "thb_mmk",           "type": "Number",  "definition": "Thai Baht to Myanmar Kyat rate", "example": "98.6", "new": False},
            {"col": "cny_mmk",           "type": "Number",  "definition": "Chinese Yuan to Myanmar Kyat rate", "example": "486", "new": False},
            {"col": "usd_mmk_change_pct","type": "Number",  "definition": "% change in USD/MMK from previous day", "example": "0.4", "new": False},
        ],
        "sample_rows": [
            ["2026-03-15", 3520, 3802, 2605, 98.6, 486, 0.4],
            ["2026-03-16", 3535, 3818, 2612, 99.1, 488, 0.43],
            ["2026-03-17", 3580, 3860, 2640, 100.2, 494, 1.27],
        ],
    },
    {
        "name": "solar_generation",
        "priority": "OPTIONAL",
        "tab_color": TAB_GREEN,
        "frequency": "Daily — hourly readings for solar-equipped sites only (export from inverter portal)",
        "columns": [
            {"col": "date",             "type": "Date",    "definition": "Generation date in YYYY-MM-DD format", "example": "2026-03-15", "new": False},
            {"col": "hour",             "type": "Number",  "definition": "Hour of day (5-18 for daylight hours)", "example": "12", "new": False},
            {"col": "store_id",         "type": "Text",    "definition": "Solar-equipped store. Must have has_solar=TRUE in stores sheet.", "example": "RH-001", "new": False},
            {"col": "solar_kwh",        "type": "Number",  "definition": "Solar energy generated this hour in kWh. From inverter monitoring.", "example": "42.5", "new": False},
            {"col": "solar_capacity_kw","type": "Number",  "definition": "Installed solar panel capacity in kW (stays same for a site)", "example": "100", "new": False},
            {"col": "efficiency_pct",   "type": "Number",  "definition": "Panel efficiency this hour as percentage (0-100). Lower on cloudy days.", "example": "85.0", "new": False},
        ],
        "sample_rows": [
            ["2026-03-15", 10, "RH-001", 32.0, 100, 72.0],
            ["2026-03-15", 11, "RH-001", 42.5, 100, 85.0],
            ["2026-03-15", 12, "RH-001", 48.0, 100, 96.0],
        ],
    },
    {
        "name": "temperature_logs",
        "priority": "OPTIONAL",
        "tab_color": TAB_GREEN,
        "frequency": "4x daily — readings at 12am, 6am, 12pm, 6pm for cold chain stores",
        "columns": [
            {"col": "date",           "type": "Date",    "definition": "Reading date in YYYY-MM-DD format", "example": "2026-03-15", "new": False},
            {"col": "hour",           "type": "Number",  "definition": "Reading hour: 0, 6, 12, or 18 (4 readings per day per zone)", "example": "6", "new": False, "dropdown": "0,6,12,18"},
            {"col": "store_id",       "type": "Text",    "definition": "Cold chain store. Must have cold_chain=TRUE in stores sheet.", "example": "RH-001", "new": False},
            {"col": "zone",           "type": "Text",    "definition": "Cold storage zone: Dairy (target 4°C), Frozen (-18°C), or Fresh Produce (6°C)", "example": "Dairy", "new": False, "dropdown": "Dairy,Frozen,Fresh Produce"},
            {"col": "temperature_c",  "type": "Number",  "definition": "Actual temperature reading in Celsius from thermometer/sensor", "example": "4.2", "new": False},
            {"col": "target_temp_c",  "type": "Number",  "definition": "Target temperature for this zone. Dairy=4, Frozen=-18, Fresh=6", "example": "4.0", "new": False},
            {"col": "critical_high_c","type": "Number",  "definition": "Maximum safe temperature. Above this = spoilage risk. Dairy=8, Frozen=-12, Fresh=10", "example": "8.0", "new": False},
            {"col": "critical_low_c", "type": "Number",  "definition": "Minimum safe temperature. Below this = damage risk. Dairy=0, Frozen=-25, Fresh=2", "example": "0.0", "new": False},
            {"col": "is_breach",      "type": "Boolean", "definition": "Was temperature outside safe range (above critical_high or below critical_low)?", "example": "FALSE", "new": False, "dropdown": "TRUE,FALSE"},
        ],
        "sample_rows": [
            ["2026-03-15", 6, "RH-001", "Dairy", 4.2, 4.0, 8.0, 0.0, "FALSE"],
            ["2026-03-15", 6, "RH-001", "Frozen", -17.5, -18.0, -12.0, -25.0, "FALSE"],
            ["2026-03-15", 12, "RH-001", "Dairy", 9.1, 4.0, 8.0, 0.0, "TRUE"],
        ],
    },
    # ── NEW: Procurement & Logistics Sheets ──
    {
        "name": "supplier_master",
        "priority": "PROCUREMENT",
        "tab_color": TAB_PURPLE,
        "frequency": "One-time setup — update quarterly or when suppliers change",
        "columns": [
            {"col": "supplier_id",        "type": "Text",   "definition": "Unique supplier identifier (e.g. SUP-001). Used in all procurement records.", "example": "SUP-001", "new": True},
            {"col": "supplier_name",      "type": "Text",   "definition": "Official supplier/company name", "example": "Myanmar Petroleum Corp", "new": True},
            {"col": "region",             "type": "Text",   "definition": "Supplier's operating region or depot location", "example": "Yangon Central", "new": True},
            {"col": "contact_person",     "type": "Text",   "definition": "Primary contact name at supplier", "example": "U Kyaw Zin", "new": True},
            {"col": "contact_phone",      "type": "Text",   "definition": "Contact phone number", "example": "+95-9-xxxx-xxxx", "new": True},
            {"col": "avg_lead_time_days", "type": "Number", "definition": "Average delivery lead time in days (historical)", "example": "1.5", "new": True},
            {"col": "price_markup_pct",   "type": "Number", "definition": "Supplier's typical price markup/discount vs market price. +2 = 2% above market, -1 = 1% below market (bulk discount)", "example": "2.0", "new": True},
            {"col": "min_order_liters",   "type": "Number", "definition": "Minimum order quantity in liters. Smaller orders may have higher markup.", "example": "200", "new": True},
            {"col": "bulk_discount_pct",  "type": "Number", "definition": "Extra discount % for orders above bulk_threshold_liters. e.g. 1.5 = 1.5% off for bulk orders", "example": "1.5", "new": True},
            {"col": "bulk_threshold_liters","type": "Number","definition": "Order quantity above which bulk_discount_pct applies", "example": "1000", "new": True},
            {"col": "payment_terms",      "type": "Text",   "definition": "Payment terms: COD (cash on delivery), NET7 (7 days), NET14, NET30, PREPAID", "example": "NET7", "new": True, "dropdown": "COD,NET7,NET14,NET30,PREPAID"},
            {"col": "reliability_rating", "type": "Text",   "definition": "Internal rating based on delivery performance: A (>95% on-time), B (85-95%), C (70-85%), D (<70%)", "example": "A", "new": True, "dropdown": "A,B,C,D"},
            {"col": "serves_sectors",     "type": "Text",   "definition": "Which sectors this supplier serves. Comma-separated: Retail,F&B,Distribution,Property or ALL", "example": "ALL", "new": True},
            {"col": "emergency_available", "type": "Boolean","definition": "Can this supplier do emergency same-day delivery? TRUE/FALSE", "example": "TRUE", "new": True, "dropdown": "TRUE,FALSE"},
            {"col": "notes",              "type": "Text",   "definition": "Any notes about this supplier (quality issues, preferred for certain sites, etc.)", "example": "Reliable but slow during rainy season", "new": True},
        ],
        "sample_rows": [
            ["SUP-001", "Myanmar Petroleum Corp", "Yangon Central", "U Kyaw Zin", "+95-9-1234-5678", 1.5, 2.0, 200, 1.5, 1000, "NET7", "A", "ALL", "TRUE", "Primary supplier, best reliability"],
            ["SUP-002", "Shwe Taung Fuel Trading", "Yangon South", "Ma Aye Aye", "+95-9-2345-6789", 2.0, -1.0, 500, 2.0, 2000, "NET14", "B", "Retail,Distribution", "TRUE", "Bulk discount specialist"],
            ["SUP-003", "Golden Eagle Diesel", "Mandalay", "Ko Aung Thu", "+95-9-3456-7890", 3.5, 5.0, 100, 0.0, 5000, "COD", "C", "ALL", "FALSE", "Emergency backup, expensive but available"],
            ["SUP-004", "Delta Fuel Supply", "Ayeyarwady", "Daw Thin Thin", "+95-9-4567-8901", 1.0, 0.0, 300, 1.0, 1500, "NET7", "A", "F&B,Property", "TRUE", "Local supplier, fast delivery in delta region"],
        ],
    },
    {
        "name": "diesel_procurement",
        "priority": "PROCUREMENT",
        "tab_color": TAB_PURPLE,
        "frequency": "Per purchase order — log every diesel purchase",
        "columns": [
            {"col": "po_number",              "type": "Text",   "definition": "Unique purchase order number. Format: PO-YYYY-NNNN", "example": "PO-2026-0342", "new": True},
            {"col": "po_date",                "type": "Date",   "definition": "Date the purchase order was placed (YYYY-MM-DD)", "example": "2026-03-14", "new": True},
            {"col": "supplier_id",            "type": "Text",   "definition": "Must match supplier_id in supplier_master tab", "example": "SUP-001", "new": True},
            {"col": "ordered_by",             "type": "Text",   "definition": "Person who placed the order (procurement team member)", "example": "U Min Htet", "new": True},
            {"col": "quantity_liters",        "type": "Number", "definition": "Total liters ordered", "example": "500", "new": True},
            {"col": "market_price_per_liter", "type": "Number", "definition": "Market reference price on order date (from diesel_prices.csv). Used to calculate markup.", "example": "4850", "new": True},
            {"col": "actual_price_per_liter", "type": "Number", "definition": "ACTUAL price paid per liter. Varies by supplier markup, volume discount, urgency premium. This is the real cost.", "example": "4947", "new": True},
            {"col": "total_cost_mmk",         "type": "Number", "definition": "Total PO cost = quantity x actual_price_per_liter (MMK)", "example": "2473500", "new": True},
            {"col": "price_vs_market_pct",    "type": "Number", "definition": "% above/below market. Positive = paid more than market. Auto-calculate: (actual - market) / market * 100", "example": "2.0", "new": True},
            {"col": "delivery_location",      "type": "Text",   "definition": "Where fuel delivers to: a store_id (e.g. RH-001) or CENTRAL-DEPOT", "example": "RH-001", "new": True},
            {"col": "promised_delivery_date", "type": "Date",   "definition": "Date supplier promised to deliver", "example": "2026-03-15", "new": True},
            {"col": "actual_delivery_date",   "type": "Date",   "definition": "Date fuel actually arrived. Blank if not yet delivered.", "example": "2026-03-15", "new": True},
            {"col": "delivery_delay_days",    "type": "Number", "definition": "Days late = actual - promised. 0 = on time, negative = early. Blank if pending.", "example": "0", "new": True},
            {"col": "status",                 "type": "Text",   "definition": "Order status: ORDERED, IN_TRANSIT, DELIVERED, CANCELLED", "example": "DELIVERED", "new": True, "dropdown": "ORDERED,IN_TRANSIT,DELIVERED,CANCELLED"},
            {"col": "order_type",             "type": "Text",   "definition": "Why was this ordered? SCHEDULED (routine), BULK_BUY (AI recommended), EMERGENCY (stockout risk), REALLOCATION (transfer)", "example": "SCHEDULED", "new": True, "dropdown": "SCHEDULED,BULK_BUY,EMERGENCY,REALLOCATION"},
            {"col": "ai_recommended",         "type": "Boolean","definition": "Was this purchase recommended by the AI COMMANDER agent? TRUE/FALSE", "example": "FALSE", "new": True, "dropdown": "TRUE,FALSE"},
            {"col": "notes",                  "type": "Text",   "definition": "Any notes (reason for emergency order, supplier negotiation, etc.)", "example": "", "new": True},
        ],
        "sample_rows": [
            ["PO-2026-0342", "2026-03-14", "SUP-001", "U Min Htet", 500, 4850, 4947, 2473500, 2.0, "RH-001", "2026-03-15", "2026-03-15", 0, "DELIVERED", "SCHEDULED", "FALSE", ""],
            ["PO-2026-0343", "2026-03-14", "SUP-002", "U Min Htet", 2000, 4850, 4802, 9604000, -1.0, "CENTRAL-DEPOT", "2026-03-16", "2026-03-16", 0, "DELIVERED", "BULK_BUY", "TRUE", "AI recommended bulk buy before price spike"],
            ["PO-2026-0344", "2026-03-15", "SUP-003", "Ma Yin Mar", 150, 4920, 5166, 774900, 5.0, "FB-001", "2026-03-15", "2026-03-15", 0, "DELIVERED", "EMERGENCY", "FALSE", "Stockout risk at FB-001, emergency same-day"],
            ["PO-2026-0345", "2026-03-16", "SUP-001", "U Min Htet", 800, 4920, 5018, 4014400, 2.0, "DW-001", "2026-03-18", "", "", "IN_TRANSIT", "SCHEDULED", "FALSE", ""],
        ],
    },
    {
        "name": "diesel_transfers",
        "priority": "PROCUREMENT",
        "tab_color": TAB_PURPLE,
        "frequency": "Per transfer event — log every fuel movement between locations",
        "columns": [
            {"col": "transfer_id",     "type": "Text",   "definition": "Unique transfer identifier. Format: TRF-YYYY-NNNN", "example": "TRF-2026-0089", "new": True},
            {"col": "transfer_date",   "type": "Date",   "definition": "Date transfer was initiated (YYYY-MM-DD)", "example": "2026-03-15", "new": True},
            {"col": "from_location",   "type": "Text",   "definition": "Source: a store_id (e.g. DW-001) or CENTRAL-DEPOT. Where fuel is coming FROM.", "example": "CENTRAL-DEPOT", "new": True},
            {"col": "to_location",     "type": "Text",   "definition": "Destination: a store_id. Where fuel is going TO.", "example": "FB-001", "new": True},
            {"col": "quantity_liters", "type": "Number", "definition": "Liters transferred", "example": "100", "new": True},
            {"col": "reason",          "type": "Text",   "definition": "Why: REALLOCATION (surplus→deficit), SCHEDULED (routine depot→site), EMERGENCY (critical stockout)", "example": "REALLOCATION", "new": True, "dropdown": "REALLOCATION,SCHEDULED,EMERGENCY"},
            {"col": "authorized_by",   "type": "Text",   "definition": "Who approved this transfer (Sector Lead or Holdings GECC for large transfers)", "example": "Sector Lead - Retail", "new": True},
            {"col": "transport_cost_mmk","type": "Number","definition": "Cost of transport/logistics for this transfer (MMK). 0 if internal vehicle.", "example": "25000", "new": True},
            {"col": "status",          "type": "Text",   "definition": "Transfer status: PENDING, IN_TRANSIT, COMPLETED, CANCELLED", "example": "COMPLETED", "new": True, "dropdown": "PENDING,IN_TRANSIT,COMPLETED,CANCELLED"},
            {"col": "linked_po",       "type": "Text",   "definition": "If transfer is from a bulk purchase, link to PO number. Blank if internal reallocation.", "example": "PO-2026-0343", "new": True},
            {"col": "notes",           "type": "Text",   "definition": "Any notes about this transfer", "example": "Urgent reallocation from surplus warehouse", "new": True},
        ],
        "sample_rows": [
            ["TRF-2026-0089", "2026-03-15", "CENTRAL-DEPOT", "FB-001", 100, "REALLOCATION", "Sector Lead - F&B", 25000, "COMPLETED", "PO-2026-0343", "Urgent reallocation from depot bulk purchase"],
            ["TRF-2026-0090", "2026-03-15", "DW-001", "RC-003", 80, "REALLOCATION", "Sector Lead - Retail", 15000, "COMPLETED", "", "DW-001 had 8 days coverage, RC-003 had 1.5 days"],
            ["TRF-2026-0091", "2026-03-16", "CENTRAL-DEPOT", "RH-001", 300, "SCHEDULED", "U Min Htet", 0, "IN_TRANSIT", "PO-2026-0343", "Routine depot to hypermarket"],
        ],
    },
    {
        "name": "generator_maint",
        "priority": "PROCUREMENT",
        "tab_color": TAB_PURPLE,
        "frequency": "Per maintenance event — log every generator service",
        "columns": [
            {"col": "maintenance_id",   "type": "Text",   "definition": "Unique maintenance record ID. Format: MNT-YYYY-NNNN", "example": "MNT-2026-0045", "new": True},
            {"col": "date",             "type": "Date",   "definition": "Date maintenance was performed (YYYY-MM-DD)", "example": "2026-03-10", "new": True},
            {"col": "store_id",         "type": "Text",   "definition": "Store where generator is located. Must match stores.csv.", "example": "RH-001", "new": True},
            {"col": "maintenance_type", "type": "Text",   "definition": "Type: SCHEDULED (routine service), EMERGENCY (breakdown), INSPECTION (check-up)", "example": "SCHEDULED", "new": True, "dropdown": "SCHEDULED,EMERGENCY,INSPECTION"},
            {"col": "description",      "type": "Text",   "definition": "What was done: oil change, filter replacement, overhaul, etc.", "example": "Oil change + air filter replacement", "new": True},
            {"col": "downtime_hours",   "type": "Number", "definition": "Hours generator was unavailable during maintenance (0 if done during grid availability)", "example": "2.0", "new": True},
            {"col": "cost_mmk",         "type": "Number", "definition": "Total maintenance cost in MMK (parts + labour)", "example": "350000", "new": True},
            {"col": "parts_replaced",   "type": "Text",   "definition": "List of parts replaced (comma-separated)", "example": "Oil filter, Air filter, Engine oil 15L", "new": True},
            {"col": "next_service_date","type": "Date",   "definition": "Recommended date for next scheduled service", "example": "2026-06-10", "new": True},
            {"col": "technician",       "type": "Text",   "definition": "Name of technician or service company", "example": "U Than Win - PowerGen Services", "new": True},
            {"col": "generator_hours_at_service","type": "Number","definition": "Generator hour meter reading at time of service. Tracks total runtime.", "example": "4520", "new": True},
            {"col": "status",           "type": "Text",   "definition": "COMPLETED, IN_PROGRESS, SCHEDULED (upcoming)", "example": "COMPLETED", "new": True, "dropdown": "COMPLETED,IN_PROGRESS,SCHEDULED"},
        ],
        "sample_rows": [
            ["MNT-2026-0045", "2026-03-10", "RH-001", "SCHEDULED", "Oil change + air filter replacement", 2.0, 350000, "Oil filter, Air filter, Engine oil 15L", "2026-06-10", "U Than Win - PowerGen Services", 4520, "COMPLETED"],
            ["MNT-2026-0046", "2026-03-12", "FB-001", "EMERGENCY", "Fuel injector failure - replaced", 8.0, 850000, "Fuel injector x2, Gasket set", "2026-04-12", "Generator Myanmar Co.", 2890, "COMPLETED"],
            ["MNT-2026-0047", "2026-03-20", "DW-001", "SCHEDULED", "500-hour service interval", 0, 0, "", "2026-03-20", "U Than Win - PowerGen Services", 5100, "SCHEDULED"],
        ],
    },
]


def _apply_cell_style(cell, font, fill, alignment=None):
    cell.font = font
    cell.fill = fill
    cell.border = THIN_BORDER
    if alignment:
        cell.alignment = alignment


def _write_readme(ws):
    """Write the README / instructions sheet."""
    ws.sheet_properties.tabColor = TAB_BLUE
    ws.column_dimensions["A"].width = 80

    instructions = [
        ("ENERGY INTELLIGENCE SYSTEM — MASTER DATA TEMPLATE", True, "1a237e"),
        ("", False, None),
        ("This Excel workbook contains templates for all 12 data files needed by the Energy Intelligence System.", False, None),
        ("Share this file with your site managers and data champions. Each tab = one data file.", False, None),
        ("", False, None),
        ("HOW TO USE THIS TEMPLATE", True, "00897b"),
        ("1. Fill in the 'stores' tab FIRST — this is your store registry (one-time setup)", False, None),
        ("2. Fill daily tabs (daily_energy, diesel_prices, diesel_inventory) EVERY DAY by 8:00 PM", False, None),
        ("3. Fill store_sales hourly data from your POS system export", False, None),
        ("4. Fill solar_generation and temperature_logs if you have those data sources", False, None),
        ("5. Upload completed tabs as CSV to the Energy Intelligence System Data Upload page", False, None),
        ("", False, None),
        ("COLUMN COLOUR GUIDE", True, "00897b"),
        ("Dark Blue header = EXISTING column (already in system)", False, None),
        ("Teal header = NEW column (added to support new features — EBITDA/hr, supplier risk, compliance)", False, None),
        ("Row 2 (grey) = Column definition — read this to understand what to enter", False, None),
        ("Row 3+ = Sample data — delete these and enter your real data", False, None),
        ("", False, None),
        ("TAB COLOUR GUIDE", True, "00897b"),
        ("RED tab = REQUIRED — system cannot function without this data", False, None),
        ("ORANGE tab = RECOMMENDED — enables important features (stockout alerts, profitability)", False, None),
        ("GREEN tab = OPTIONAL — enables advanced features (solar optimization, spoilage prediction)", False, None),
        ("PURPLE tab = PROCUREMENT & LOGISTICS — diesel purchasing, transfers, supplier management", False, None),
        ("", False, None),
        ("SUBMISSION FREQUENCY", True, "00897b"),
        ("stores — One-time setup, update when stores open/close", False, None),
        ("daily_energy — DAILY by 8:00 PM (site manager)", False, None),
        ("diesel_prices — DAILY (procurement team)", False, None),
        ("diesel_inventory — DAILY end-of-day (site manager)", False, None),
        ("store_sales — DAILY export from POS (auto or manual)", False, None),
        ("fx_rates — DAILY (finance team)", False, None),
        ("solar_generation — DAILY export from inverter portal (facilities)", False, None),
        ("temperature_logs — 4x DAILY at 12am, 6am, 12pm, 6pm (cold chain stores)", False, None),
        ("", False, None),
        ("VALIDATION RULES", True, "00897b"),
        ("• generator_hours + grid_hours must be <= 24 hours", False, None),
        ("• diesel_stock variance > 20% from yesterday triggers validation alert", False, None),
        ("• All dates must be in YYYY-MM-DD format", False, None),
        ("• store_id must match exactly across all files (case-sensitive)", False, None),
        ("• Dropdown columns have pre-set values — use only those", False, None),
        ("", False, None),
        ("NEW COLUMNS EXPLAINED", True, "00897b"),
        ("operating_mode_actual/planned — Tracks what AI recommended vs what store actually did (adoption tracking)", False, None),
        ("submitted_by / submitted_at — Tracks data submission compliance per site", False, None),
        ("regional_shortage_flag — Market supply signal for supplier risk model", False, None),
        ("labour_cost_mmk — Enables EBITDA per operating hour calculation (the core profitability engine)", False, None),
        ("", False, None),
        ("NEW TABS EXPLAINED", True, "00897b"),
        ("supplier_master — Supplier registry with pricing tiers. Each supplier has different price markup from market.", False, None),
        ("diesel_procurement — Purchase orders. Actual price paid per PO (varies by supplier, volume, urgency).", False, None),
        ("diesel_transfers — Fuel movement: depot→site or site→site reallocation during shortages.", False, None),
        ("generator_maintenance — Service history: scheduled/emergency maintenance, downtime, cost.", False, None),
        ("", False, None),
        ("DIESEL PRICING MODEL", True, "00897b"),
        ("diesel_prices.csv = MARKET reference price (benchmark, same for everyone)", False, None),
        ("supplier_master.csv = Each supplier has a price_markup_pct (e.g. SUP-001 = +2%, SUP-003 = +5%)", False, None),
        ("diesel_procurement.csv = Each PO has actual_price_per_liter (market x supplier markup x volume discount)", False, None),
        ("This lets the system compare: which supplier gives best price? When to bulk buy? Emergency premium cost?", False, None),
        ("", False, None),
        ("GENERATED BY", True, "1a237e"),
        ("Energy Intelligence System v1.0 — AI-Powered Agentic Framework for Energy BCP", False, None),
        ("Download fresh template anytime from the Data Upload page", False, None),
    ]

    for i, (text, is_heading, color) in enumerate(instructions, 1):
        cell = ws.cell(row=i, column=1, value=text)
        if is_heading and color:
            cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        elif is_heading:
            cell.font = Font(name="Calibri", bold=True, size=12)
        else:
            cell.font = Font(name="Calibri", size=11, color="333333")


def generate_template() -> bytes:
    """Generate the master data template Excel workbook and return as bytes."""
    wb = Workbook()

    # ── Sheet 1: README ──
    ws_readme = wb.active
    ws_readme.title = "READ ME FIRST"
    _write_readme(ws_readme)

    # ── Data sheets ──
    for sheet_def in SHEETS:
        ws = wb.create_sheet(title=sheet_def["name"])
        ws.sheet_properties.tabColor = sheet_def["tab_color"]
        columns = sheet_def["columns"]

        # Row 0 (hidden): Sheet metadata
        # Row 1: Headers
        for ci, col_def in enumerate(columns, 1):
            cell = ws.cell(row=1, column=ci, value=col_def["col"])
            if col_def.get("new"):
                _apply_cell_style(cell, NEW_COL_FONT, NEW_COL_FILL, Alignment(horizontal="center"))
            else:
                _apply_cell_style(cell, HEADER_FONT, HEADER_FILL, Alignment(horizontal="center"))

        # Row 2: Definitions
        for ci, col_def in enumerate(columns, 1):
            cell = ws.cell(row=2, column=ci, value=col_def["definition"])
            fill = NEW_DEF_FILL if col_def.get("new") else DEF_FILL
            _apply_cell_style(cell, DEF_FONT, fill, WRAP)

        # Row 3+: Sample data
        for ri, row_data in enumerate(sheet_def.get("sample_rows", []), 3):
            for ci, value in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=value)
                cell.font = SAMPLE_FONT
                cell.border = THIN_BORDER

        # Freeze top 2 rows
        ws.freeze_panes = "A3"

        # Auto-width columns (min 14, max 40)
        for ci, col_def in enumerate(columns, 1):
            col_letter = get_column_letter(ci)
            def_len = len(col_def["definition"])
            header_len = len(col_def["col"])
            width = min(40, max(14, header_len + 4, def_len // 3))
            ws.column_dimensions[col_letter].width = width

        # Data validation dropdowns (apply to rows 3-1000)
        for ci, col_def in enumerate(columns, 1):
            if "dropdown" in col_def:
                col_letter = get_column_letter(ci)
                dv = DataValidation(
                    type="list",
                    formula1=f'"{col_def["dropdown"]}"',
                    allow_blank=True,
                    showErrorMessage=True,
                    errorTitle="Invalid value",
                    error=f"Please select from: {col_def['dropdown']}",
                )
                dv.add(f"{col_letter}3:{col_letter}1000")
                ws.add_data_validation(dv)

        # Add frequency + priority note in a merged cell above header
        # (We skip this to keep it clean — the README covers it)

    # ── Write to bytes ──
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
